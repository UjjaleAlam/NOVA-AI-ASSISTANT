from openpyxl import load_workbook


def extract_text(path):
    """
    Extract text from an Excel workbook.
    """

    try:

        workbook = load_workbook(
            path,
            data_only=True
        )

        text = []

        for sheet in workbook.worksheets:

            for row in sheet.iter_rows(values_only=True):

                values = [

                    str(cell)

                    for cell in row

                    if cell is not None
                ]

                if values:

                    text.append(" ".join(values))

        workbook.close()

        return "\n".join(text)

    except Exception as e:

        print(f"[XLSX Reader] {e}")

        return ""